#!/usr/bin/env python3
"""导入订单表格：读 xlsx → 更新 orders.json（含图片提取保存）

支持两种图片：
  1. 浮动图片（Excel 插入→图片，锚定在单元格上）
  2. WPS「嵌入单元格图片」（DISPIMG 函数，存 xl/cellimages.xml + xl/media/）

用法：python import_orders.py 订单表格.xlsx
"""
import json
import os
import re
import sys
import uuid
import zipfile
from datetime import datetime, date, timedelta
from xml.etree import ElementTree as ET
from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor

BASE = '/home/ubuntu/order-tracker'
ORDERS_FILE = os.path.join(BASE, 'orders.json')
IMAGES_DIR = os.path.join(BASE, 'images')

STAGE_MAP = {
    '设计确认': 'design',
    '面料采购': 'fabric',
    '生产中': 'production',
    '生产完成': 'production',
    '已发货': 'shipping',
}


def map_stage(text):
    """阶段文本 → stage_key。标准名直接映射，否则按关键词推断。"""
    if not text:
        return 'design'
    if text in STAGE_MAP:
        return STAGE_MAP[text]
    if any(k in text for k in ('发货', '物流', '送达', '签收')):
        return 'shipping'
    if any(k in text for k in ('面料', '采购', '买布')):
        return 'fabric'
    if any(k in text for k in ('印', '生产', '裁剪', '缝', '绣', '烫画', '包装')):
        return 'production'
    return 'design'

_XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
_A = 'http://schemas.openxmlformats.org/drawingml/2006/main'
_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'


def fmt_date(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return f'{v.year}年{v.month}月{v.day}日'
    if isinstance(v, date):
        return f'{v.year}年{v.month}月{v.day}日'
    if isinstance(v, (int, float)):
        # Excel 日期序列号（46269 → 2026年9月4日）
        try:
            d = datetime(1899, 12, 30) + timedelta(days=int(v))
            return f'{d.year}年{d.month}月{d.day}日'
        except Exception:
            return str(v)
    return str(v).strip()


def _img_ext(data, fmt):
    fmt = (fmt or '').lower()
    if fmt in ('jpg', 'jpeg'):
        return '.jpg'
    if fmt in ('gif', 'webp', 'bmp'):
        return '.' + fmt
    if data[:3] == b'\xff\xd8\xff':
        return '.jpg'
    return '.png'


def _save_image(data):
    """保存图片 bytes 到 images/，返回 URL"""
    os.makedirs(IMAGES_DIR, exist_ok=True)
    ext = _img_ext(data, '')
    fname = uuid.uuid4().hex + ext
    with open(os.path.join(IMAGES_DIR, fname), 'wb') as f:
        f.write(data)
    return '/images/' + fname


def extract_floating_images(ws):
    """提取浮动图片，返回 {0-based row: '/images/xxx'}"""
    result = {}
    for img in (getattr(ws, '_images', []) or []):
        row = None
        anchor = getattr(img, 'anchor', None)
        if isinstance(anchor, (OneCellAnchor, TwoCellAnchor)):
            frm = getattr(anchor, '_from', None)
            row = getattr(frm, 'row', None)
        elif anchor is not None:
            frm = getattr(anchor, '_from', None)
            row = getattr(frm, 'row', None)
        if row is None:
            continue
        data = img._data()
        if data:
            result[int(row)] = _save_image(data)
    return result


def parse_dispimg_images(xlsx_path):
    """解析 WPS DISPIMG 单元格图片，返回 {图片ID: 图片bytes}"""
    result = {}
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            if 'xl/cellimages.xml' not in z.namelist():
                return result
            root = ET.fromstring(z.read('xl/cellimages.xml'))
            id_to_embed = {}
            for pic in root.iter(f'{{{_XDR}}}pic'):
                cnvpr = pic.find(f'{{{_XDR}}}nvPicPr/{{{_XDR}}}cNvPr')
                blip = pic.find(f'{{{_XDR}}}blipFill/{{{_A}}}blip')
                if cnvpr is None or blip is None:
                    continue
                name = cnvpr.get('name')
                embed = blip.get(f'{{{_R}}}embed')
                if name and embed:
                    id_to_embed[name] = embed
            rel_root = ET.fromstring(z.read('xl/_rels/cellimages.xml.rels'))
            embed_to_media = {rel.get('Id'): rel.get('Target') for rel in rel_root}
            for name, embed in id_to_embed.items():
                target = embed_to_media.get(embed)
                if not target:
                    continue
                media_path = target if target.startswith('xl/') else 'xl/' + target
                if media_path in z.namelist():
                    result[name] = z.read(media_path)
    except Exception as e:
        print(f'[warn] DISPIMG 图片解析失败: {e}')
    return result


def _is_empty(vals):
    return all(v is None or str(v).strip() == '' for v in vals)


def parse(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb['订单录入']
    floating = extract_floating_images(ws)
    dispimg = parse_dispimg_images(xlsx_path)

    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    orders = {}
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        if _is_empty(vals):
            continue
        r = row[0].row - 1  # 0-based，匹配浮动图片锚点
        order_id = str(vals[col['订单编号']] or '').strip()
        if not order_id or '示例' in order_id:
            continue
        name = str(vals[col['款式名称']] or '').strip()
        if not name:
            continue
        try:
            qty = int(vals[col['数量']] or 0)
        except (TypeError, ValueError):
            qty = 0
        try:
            price = float(vals[col['单价(元)']] or 0)
        except (TypeError, ValueError):
            price = 0
        if isinstance(price, float) and price == int(price):
            price = int(price)
        amount = round(qty * price, 2) if (qty and price) else 0
        if isinstance(amount, float) and amount == int(amount):
            amount = int(amount)

        # 图片：优先 WPS DISPIMG，其次浮动图片
        image = ''
        img_val = vals[col['款式图片']] if col['款式图片'] < len(vals) else None
        if isinstance(img_val, str) and 'DISPIMG' in img_val:
            m = re.search(r'DISPIMG\("([^"]+)"', img_val)
            if m:
                data = dispimg.get(m.group(1))
                if data:
                    image = _save_image(data)
        if not image:
            image = floating.get(r, '')

        note = str(vals[col['备注']] or '').strip()

        if order_id not in orders:
            orders[order_id] = {
                'customer': str(vals[col['客户名称']] or '').strip(),
                'estimated_delivery': fmt_date(vals[col['预计交期']]),
                'payment': str(vals[col['支付方式']] or '').strip(),
                'items': [],
            }
        orders[order_id]['items'].append({
            'name': name,
            'qty': qty,
            'unit_price': price,
            'amount': amount,
            'image': image,
            'note': note,
        })

    updates_map = {}
    if '进度更新' in wb.sheetnames:
        ws2 = wb['进度更新']
        headers2 = [c.value for c in ws2[1]]
        col2 = {h: i for i, h in enumerate(headers2)}
        for row in ws2.iter_rows(min_row=2):
            vals = [c.value for c in row]
            if _is_empty(vals):
                continue
            order_id = str(vals[col2['订单编号']] or '').strip()
            if not order_id or '示例' in order_id:
                continue
            stage = str(vals[col2['阶段']] or '').strip()
            note = str(vals[col2['备注']] or '').strip()
            # 若阶段列填了非标准文本（业务员常把进展描述写这里），当作备注
            final_note = note or (stage if stage not in STAGE_MAP else '')
            updates_map.setdefault(order_id, []).append({
                'date': fmt_date(vals[col2['更新日期']]),
                'stage_key': map_stage(stage),
                'note': final_note or stage,
            })

    return orders, updates_map


def apply(xlsx_path, delete_missing=False):
    orders, updates_map = parse(xlsx_path)
    with open(ORDERS_FILE, encoding='utf-8') as f:
        data = json.load(f)

    new_count = update_count = 0
    for order_id, o in orders.items():
        total = sum(it['amount'] for it in o['items'])
        if isinstance(total, float) and total == int(total):
            total = int(total)
        existing = data['orders'].get(order_id)
        if existing:
            existing['customer'] = o['customer'] or existing.get('customer', '')
            existing['estimated_delivery'] = o['estimated_delivery'] or existing.get('estimated_delivery', '')
            existing['payment'] = o['payment'] or existing.get('payment', '')
            existing['items'] = o['items']
            if total:
                existing['amount'] = '¥' + str(total)
            update_count += 1
        else:
            data['orders'][order_id] = {
                'customer': o['customer'],
                'product': '',
                'stage_key': 'design',
                'estimated_delivery': o['estimated_delivery'],
                'amount': ('¥' + str(total)) if total else '',
                'payment': o['payment'],
                'items': o['items'],
                'updates': [],
                'tracking_number': '',
                'carrier': '',
                'tracking_history': [],
            }
            new_count += 1

    for order_id, ups in updates_map.items():
        if order_id not in data['orders']:
            continue
        order = data['orders'][order_id]
        existing_notes = {(u['date'], u['note']) for u in order.get('updates', [])}
        for u in ups:
            if (u['date'], u['note']) not in existing_notes:
                order.setdefault('updates', []).append(u)
                existing_notes.add((u['date'], u['note']))
        if order.get('updates'):
            order['stage_key'] = order['updates'][-1]['stage_key']

    # 完全同步：删除表格中不存在的订单（连同其图片）
    deleted = []
    if delete_missing:
        table_ids = set(orders.keys())
        for oid in list(data['orders'].keys()):
            if oid not in table_ids:
                order = data['orders'].pop(oid)
                for it in order.get('items', []):
                    img = it.get('image', '')
                    if img:
                        f = os.path.join(IMAGES_DIR, os.path.basename(img))
                        if os.path.exists(f):
                            os.remove(f)
                deleted.append(oid)

    with open(ORDERS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    return new_count, update_count, len(updates_map), deleted


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('xlsx', nargs='?', default=os.path.join(BASE, '订单录入模板.xlsx'), help='订单表格路径')
    parser.add_argument('--sync', action='store_true', help='完全同步：删除表格中不存在的订单')
    args = parser.parse_args()
    n, u, p, deleted = apply(args.xlsx, delete_missing=args.sync)
    print(f'导入完成：新增 {n} 个订单，更新 {u} 个订单，{p} 个订单有进度更新')
    if deleted:
        print(f'已同步删除 {len(deleted)} 个订单：{", ".join(deleted)}')
    else:
        print('（本次未删除订单；加 --sync 会完全同步删除表格中不存在的订单）')
