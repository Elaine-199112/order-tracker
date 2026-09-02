#!/usr/bin/env python3
"""生成订单录入 Excel 模板（供业务员填写，每日发回导入）"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='0D1F2D')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
CENTER = Alignment(horizontal='center', vertical='center')

wb = Workbook()

# ── Sheet1 订单录入（每个款式一行）──
ws = wb.active
ws.title = '订单录入'
headers = ['订单编号', '客户名称', '预计交期', '支付方式', '款式名称', '数量', '单价(元)', '款式图片', '备注']
ws.append(headers)
for col in range(1, len(headers) + 1):
    c = ws.cell(row=1, column=col)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
for i, w in enumerate([18, 16, 12, 14, 28, 8, 10, 24, 22], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 示例行（浅色，业务员照着填，填完删）
example = ['示例', '示例客户', '9月9日', '已支付全款', '【示例-可删除】紫色数码印冰丝短袖T恤', 200, 36, '(图片插这里)', '']
ws.append(example)
for col in range(1, len(headers) + 1):
    ws.cell(row=2, column=col).fill = PatternFill('solid', fgColor='F5EBE0')
ws.row_dimensions[2].height = 60

# ── Sheet2 进度更新（每次进度一行）──
ws2 = wb.create_sheet('进度更新')
headers2 = ['订单编号', '更新日期', '阶段', '备注']
ws2.append(headers2)
for col in range(1, len(headers2) + 1):
    c = ws2.cell(row=1, column=col)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
for i, w in enumerate([18, 12, 16, 34], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ── Sheet3 使用说明 ──
ws3 = wb.create_sheet('使用说明')
ws3.column_dimensions['A'].width = 90
lines = [
    ('【怎么填订单】', True),
    ('1. 在「订单录入」表里，每个款式填一行。', False),
    ('2. 同一个订单有多个款式时，用相同的「订单编号」（客户/交期/支付方式只填第一行即可）。', False),
    ('3. 「数量」「单价」填数字，单价单位是元。', False),
    ('4. 款式图片：点「款式图片」列的单元格 → 菜单「插入 → 图片」→ 选图片插入。', False),
    ('5. 每天填完保存，把整个表格发给老板。', False),
    ('', False),
    ('【怎么填进度】', True),
    ('1. 在「进度更新」表里，每次进度有变化就填一行。', False),
    ('2. 「阶段」填：设计确认 / 面料采购 / 生产中 / 生产完成 / 已发货。', False),
    ('3. 「备注」写具体进展（如：印花已打完、等待裁剪）。', False),
    ('', False),
    ('【注意】', True),
    ('「订单录入」表第2行是示例，照着格式填，填完删掉即可。', False),
]
for i, (text, bold) in enumerate(lines, 1):
    ws3.cell(row=i, column=1, value=text).font = Font(bold=bold, size=12 if bold else 11)

out = '/home/ubuntu/order-tracker/订单录入模板.xlsx'
wb.save(out)
print('模板已生成:', out)
